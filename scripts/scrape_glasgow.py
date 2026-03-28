"""
PubMed scraper for Glasgow University researchers.

Reads researcher names from the MVLS Imaging Initiative xlsx file,
searches PubMed for each researcher's publications, and fetches all abstracts.

Outputs:
  data/glasgow_abstracts.csv  — one row per unique paper (pmid, year, journal, title, abstract, …)
  data/glasgow_authors.csv    — author–paper mapping (pmid, author_name, school, college)

Usage:
    uv run python scripts/scrape_glasgow.py
"""

from Bio import Entrez
import pandas as pd
import openpyxl
from tqdm import tqdm
import time
import os
import json
import re

# -- Configuration --
Entrez.email = "christoph.daube@gmail.com"
Entrez.api_key = "83596db810992244490dd34f950166198307"

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
XLSX_FILE = os.path.join(DATA_DIR, "MVLS Imaging Initiative List of Academics.xlsx")
OUTPUT_CSV = os.path.join(DATA_DIR, "glasgow_abstracts.csv")
AUTHORS_CSV = os.path.join(DATA_DIR, "glasgow_authors.csv")
PMID_CACHE = os.path.join(DATA_DIR, ".glasgow_pmid_cache.json")
FETCH_BATCH = 200
MAX_RETRIES = 5
SAVE_EVERY = 5000


# ---------------------------------------------------------------------------
# 1. Parse the xlsx
# ---------------------------------------------------------------------------

def extract_name(text):
    """Extract a person's name from 'Name <email>' or 'first.last@email' formats."""
    # "Name (affiliation) <email>" or "Name <email>"
    m = re.match(r"^([^<(]+?)(?:\s*\(.*?\))?\s*<", text)
    if m:
        return m.group(1).strip()
    # plain email "first.last@domain"
    m = re.match(r"^([A-Za-z]+)\.([A-Za-z]+)", text)
    if m:
        return f"{m.group(1).title()} {m.group(2).title()}"
    return None


def parse_researchers(xlsx_path):
    """Return a list of dicts with keys: name, school, college."""
    wb = openpyxl.load_workbook(xlsx_path)
    researchers = []
    seen = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            vals = list(row)
            col1 = str(vals[0] or "").strip()
            if not col1:
                continue

            name = extract_name(col1)
            if not name:
                continue

            if sheet_name == "NHS":
                # NHS sheet: col2 = Glasgow email, col3 = school abbreviation
                school = str(vals[2] or "").strip() if len(vals) > 2 and vals[2] else "NHS"
                college = "NHS"
            else:
                school = str(vals[1] or "").strip() if len(vals) > 1 and vals[1] else sheet_name
                college = str(vals[2] or "").strip() if len(vals) > 2 and vals[2] else sheet_name

            key = (name, school, college)
            if key in seen:
                continue
            seen.add(key)

            researchers.append({"name": name, "school": school, "college": college})

    return researchers


# ---------------------------------------------------------------------------
# 2. PubMed search helpers (mirrors scrape_pubmed.py patterns)
# ---------------------------------------------------------------------------

# Affiliation terms to filter for Glasgow-based researchers
GLASGOW_AFFIL_TERMS = [
    "Glasgow", "Glas", "GGC",  # covers University of Glasgow, NHS GGC, etc.
]

GLASGOW_AFFIL_FILTER = " OR ".join(f'"{t}"[Affiliation]' for t in GLASGOW_AFFIL_TERMS)


def author_query(name):
    """Build a PubMed author+affiliation search term from a full name."""
    parts = name.split()
    if len(parts) >= 2:
        last = parts[-1]
        first = parts[0]
        author_part = f'"{last} {first}"[Author]'
    else:
        author_part = f'"{name}"[Author]'
    return f'{author_part} AND ({GLASGOW_AFFIL_FILTER})'


def get_pmids_for_author(name):
    """Return all PMIDs for a given author name."""
    query = author_query(name)
    for attempt in range(MAX_RETRIES):
        try:
            handle = Entrez.esearch(db="pubmed", term=query, retmax=10000)
            results = Entrez.read(handle)
            handle.close()
            count = int(results["Count"])
            pmids = results["IdList"]
            if count > 10000:
                tqdm.write(f"  WARNING: {name} has {count} results (capped at 10 000)")
            return pmids
        except Exception as e:
            wait = 2 ** (attempt + 1)
            tqdm.write(f"  Search error for {name}: {e}. Retry in {wait}s…")
            time.sleep(wait)
    return []


def extract_article_data(article):
    """Pull fields from a single PubMed XML record."""
    try:
        medline = article["MedlineCitation"]
        art = medline["Article"]

        title = str(art.get("ArticleTitle", ""))

        abstract_parts = art.get("Abstract", {}).get("AbstractText", [])
        if not abstract_parts:
            return None
        abstract = " ".join(str(p) for p in abstract_parts)

        pub_date = art["Journal"]["JournalIssue"]["PubDate"]
        year = pub_date.get("Year") or pub_date.get("MedlineDate", "Unknown")[:4]

        journal = art["Journal"].get("Title", "Unknown")
        pmid = str(medline["PMID"])

        mesh_list = medline.get("MeshHeadingList", [])
        mesh_str = "; ".join(str(mh.get("DescriptorName", "")) for mh in mesh_list)

        doi = ""
        for eid in art.get("ELocationID", []):
            if str(eid.attributes.get("EIdType", "")) == "doi":
                doi = str(eid)
                break

        # Full author list + affiliations for reference
        author_list = art.get("AuthorList", [])
        authors = []
        affiliations_all = []
        for au in author_list:
            last = au.get("LastName", "")
            fore = au.get("ForeName", "")
            if last:
                authors.append(f"{last} {fore}".strip())
            for aff in au.get("AffiliationInfo", []):
                affiliations_all.append(str(aff.get("Affiliation", "")))

        return {
            "pmid": pmid,
            "year": year,
            "journal": journal,
            "title": title,
            "abstract": abstract,
            "mesh_terms": mesh_str,
            "doi": doi,
            "all_authors": "; ".join(authors),
            "_affiliations": affiliations_all,  # used for post-filtering
        }
    except (KeyError, TypeError, IndexError, AttributeError):
        return None


def _has_glasgow_affiliation(affiliations):
    """Check if any affiliation string mentions Glasgow."""
    for aff in affiliations:
        aff_lower = aff.lower()
        if "glasgow" in aff_lower:
            return True
    return False


def fetch_articles_by_pmid(pmids):
    """Fetch article details for a batch of PMIDs, filtering for Glasgow affiliations."""
    id_str = ",".join(pmids)
    handle = Entrez.efetch(db="pubmed", id=id_str, rettype="abstract", retmode="xml")
    records = Entrez.read(handle)
    handle.close()

    results = []
    for article in records.get("PubmedArticle", []):
        row = extract_article_data(article)
        if row and _has_glasgow_affiliation(row.pop("_affiliations", [])):
            results.append(row)
    return results


# ---------------------------------------------------------------------------
# 3. Main pipeline
# ---------------------------------------------------------------------------

def scrape():
    # ── Parse researchers ──
    print("Parsing researchers from xlsx…")
    researchers = parse_researchers(XLSX_FILE)
    colleges = sorted(set(r["college"] for r in researchers))
    print(f"  {len(researchers)} researchers across colleges: {', '.join(colleges)}")

    # ── Search PubMed per author ──
    if os.path.exists(PMID_CACHE):
        with open(PMID_CACHE) as f:
            author_pmids = json.load(f)
        print(f"  Loaded cached PMIDs for {len(author_pmids)} researchers.")
    else:
        author_pmids = {}
        print("Searching PubMed for each researcher…")
        for r in tqdm(researchers, desc="Author search"):
            name = r["name"]
            if name in author_pmids:
                continue
            author_pmids[name] = get_pmids_for_author(name)
            time.sleep(0.11)

        with open(PMID_CACHE, "w") as f:
            json.dump(author_pmids, f)

    # ── Build author-paper mapping ──
    author_rows = []
    all_pmids = set()
    for r in researchers:
        for pmid in author_pmids.get(r["name"], []):
            author_rows.append({
                "pmid": pmid,
                "author_name": r["name"],
                "school": r["school"],
                "college": r["college"],
            })
            all_pmids.add(pmid)

    pd.DataFrame(author_rows).to_csv(AUTHORS_CSV, index=False)
    print(f"  Author mapping ({len(author_rows)} rows) saved to {AUTHORS_CSV}")
    print(f"  {len(all_pmids)} unique PMIDs to fetch")

    # ── Fetch articles ──
    already_fetched = set()
    all_data = []
    if os.path.exists(OUTPUT_CSV):
        existing = pd.read_csv(OUTPUT_CSV)
        already_fetched = set(existing["pmid"].astype(str))
        all_data = existing.to_dict("records")
        print(f"  Resuming: {len(already_fetched)} articles already fetched.")

    remaining = [p for p in all_pmids if p not in already_fetched]
    print(f"  Remaining to fetch: {len(remaining)}")

    if not remaining:
        print("All articles already fetched!")
        return

    batches = [remaining[i : i + FETCH_BATCH] for i in range(0, len(remaining), FETCH_BATCH)]

    pbar = tqdm(batches, desc="Fetching articles")
    for i, batch in enumerate(pbar):
        for attempt in range(MAX_RETRIES):
            try:
                articles = fetch_articles_by_pmid(batch)
                all_data.extend(articles)
                pbar.set_postfix(total=len(all_data))
                break
            except Exception as e:
                wait = 2 ** (attempt + 1)
                tqdm.write(f"  Error: {e}. Retrying in {wait}s…")
                time.sleep(wait)
                if attempt == MAX_RETRIES - 1:
                    tqdm.write(f"  FAILED batch {i}. Skipping.")

        if (i + 1) % (SAVE_EVERY // FETCH_BATCH) == 0:
            df = pd.DataFrame(all_data)
            df.drop_duplicates(subset="pmid", inplace=True)
            df.to_csv(OUTPUT_CSV, index=False)
            tqdm.write(f"  Checkpoint: {len(df)} unique abstracts saved.")

    # ── Final save ──
    df = pd.DataFrame(all_data)
    df.drop_duplicates(subset="pmid", inplace=True)
    df.to_csv(OUTPUT_CSV, index=False)
    print(f"\nDone! {len(df)} unique abstracts saved to {OUTPUT_CSV}")

    if os.path.exists(PMID_CACHE):
        os.remove(PMID_CACHE)


if __name__ == "__main__":
    scrape()
